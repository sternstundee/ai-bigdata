# 导入依赖库和学生字典（替换XYY为自己的班号+学号）
from openpyxl import Workbook
from student_136 import dict_stu  # 导入student_XYY.py中的字典

# 1. 创建工作簿和工作表
wb = Workbook()  # 新建工作簿
ws = wb.create_sheet(title="student", index=0)  # 创建名为"student"的工作表（设为第一个）

# 2. 写入表头（A列：学号，B列：姓名）
ws['A1'] = "36"
ws['B1'] = "舒文璨"

# 3. 遍历字典，写入学生信息（从第2行开始）
row_num = 2  # 起始行号（跳过表头）
for student_id, student_name in dict_stu.items():
    ws[f'A{row_num}'] = student_id  # A列写学号
    ws[f'B{row_num}'] = student_name  # B列写姓名
    row_num += 1

# 4. 保存Excel文件（替换XYY为自己的班号+学号）
excel_filename = "student_136.xlsx"
wb.save(excel_filename)
print(f"✅ 学生名单已成功存入 {excel_filename}！")