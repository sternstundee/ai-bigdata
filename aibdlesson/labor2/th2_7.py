from openpyxl import load_workbook
import random
import time
import sys


def random_call_from_excel():
    # 1. 加载Excel文件（替换XYY为自己的班号+学号）
    excel_filename = "student_136.xlsx"
    try:
        wb = load_workbook(excel_filename)
        ws = wb["student"]  # 选择工作表
    except FileNotFoundError:
        print(f"❌ 未找到文件 {excel_filename}，请先运行实例6生成文件！")
        return
    except KeyError:
        print(f"❌ 工作表 'student' 不存在，请检查Excel文件结构！")
        return

    # 2. 读取Excel中的学号和姓名（跳过表头，从第2行开始）
    student_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # values_only=True：只取单元格值
        student_id, student_name = row[0], row[1]
        if student_id and student_name:  # 跳过空行
            student_list.append((student_id, student_name))

    total_students = len(student_list)
    if total_students == 0:
        print("❌ Excel中未读取到学生信息！")
        return

    # 3. 输入点名人数（1~总人数）
    while True:
        try:
            call_count = int(input(f"\n请输入点名人数（不大于{total_students}，不小于1）："))
            if 1 <= call_count <= total_students:
                break
            print(f"❌ 点名人数需在1~{total_students}之间，请重新输入！")
        except ValueError:
            print("❌ 输入错误！请输入整数。")

    # 4. 随机抽取学生并输出
    selected_students = random.sample(student_list, call_count)
    print("\n📢 随机点名结果：")
    for idx, (sid, sname) in enumerate(selected_students, 1):
        print(f"{idx:02d}: {sid:02d}→{sname}", end="  " if idx % 3 != 0 else "\n")  # 每3个换行
    print(f"\n⏰ 请上面{call_count}个同学30秒钟之内在QQ群回复1！计时开始...")

    # 5. 30秒计时（进度条形式）
    total_seconds = 30
    for remaining in range(total_seconds, 0, -1):
        # 实时更新进度条（覆盖上一行）
        sys.stdout.write(f"\r倒计时：{remaining:2d}秒 | " + "▓" * (total_seconds - remaining) + "░" * remaining)
        sys.stdout.flush()  # 强制刷新输出
        time.sleep(1)  # 暂停1秒

    # 6. 计时结束
    sys.stdout.write(f"\r倒计时：0秒 | " + "▓" * total_seconds + " | 计时结束！\n")
    print("👋 点名结束，未回复的同学按旷课处理。")


# 启动点名程序
random_call_from_excel()