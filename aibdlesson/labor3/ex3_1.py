import pymysql
from openpyxl import load_workbook

# ===================== 1. 配置MySQL连接信息（需根据自身情况修改密码）=====================
mysql_config = {
    "host": "localhost",       # 本地MySQL
    "user": "root",            # 用户名
    "password": "123456",      # 默认密码，若修改过请替换
    "port": 3306,              # MySQL默认端口
    "charset": "utf8mb4"       # 支持中文编码
}

# ===================== 2. 连接MySQL，创建数据库和表 =====================
try:
    # 步骤1：连接MySQL（暂不指定数据库，用于创建数据库）
    conn = pymysql.connect(**mysql_config)
    cursor = conn.cursor()

    # 步骤2：创建数据库 test01（班号1+学号36），不存在则创建
    db_name = "test01"
    cursor.execute(f"CREATE DATABASE IF NOT EXISTS {db_name};")
    print(f"✅ 数据库 {db_name} 创建成功（若已存在则跳过）")

    # 步骤3：切换到 student_136 数据库
    cursor.execute(f"USE {db_name};")

    # 步骤4：创建 mingdan 表（学号+姓名），不存在则创建
    create_table_sql = """
    CREATE TABLE IF NOT EXISTS mingdan (
        id INT PRIMARY KEY,       # 学号（主键，唯一）
        name VARCHAR(50) NOT NULL # 姓名（非空）
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """
    cursor.execute(create_table_sql)
    print(f"✅ 表 mingdan 创建成功（若已存在则跳过）")

    # ===================== 3. 读取Excel文件（student_136.xlsx）=====================
    excel_file = "student_136.xlsx"  # 实验二生成的Excel文件
    try:
        wb = load_workbook(excel_file)
        ws = wb["student"]  # 工作表名称（实验二定义为"student"）
        print(f"✅ 成功读取Excel文件：{excel_file}")
    except FileNotFoundError:
        print(f"❌ 未找到Excel文件 {excel_file}，请先运行实验二实例6生成该文件！")
        exit()

    # ===================== 4. 读取Excel数据并插入MySQL =====================
    student_dict = {}  # 存储学生信息的字典（key：学号，value：姓名）
    insert_sql = "INSERT IGNORE INTO mingdan (id, name) VALUES (%s, %s);"  # IGNORE避免重复插入

    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头（第1行）
        student_id, student_name = row[0], row[1]
        if student_id and student_name:  # 跳过空行
            student_dict[student_id] = student_name  # 存入字典
            cursor.execute(insert_sql, (student_id, student_name))  # 插入数据库

    conn.commit()  # 提交事务
    print(f"✅ 共插入 {len(student_dict)} 名学生信息到数据库")

    # ===================== 5. 字典输出学生信息 =====================
    print("\n📋 学生信息字典（学号: 姓名）：")
    for sid, sname in sorted(student_dict.items()):  # 按学号排序输出
        print(f"{sid:2d}: {sname}")

except pymysql.Error as e:
    print(f"❌ MySQL操作失败：{e}")
    conn.rollback()  # 出错回滚事务
finally:
    # 关闭游标和连接
    if cursor:
        cursor.close()
    if conn:
        conn.close()
    print("\n🔌 MySQL连接已关闭")